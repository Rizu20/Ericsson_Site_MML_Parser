import os
import sys
from time import sleep
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill, Side, Border
from colorama import init,Fore,Style

def parser_3G(files):
    headers=['Site Name','nodeIpAddress','ipAddress','IpInterface-1','IpInterface-2']
    
    wb=Workbook()
    ws=wb.active

    for i,j in zip(range(1,len(headers)+1),headers):
        ws.cell(row=1,column=i).value=j

    site_count=0

    for file in files:
        site_count=site_count+1
        ws.cell(row=site_count+1,column=headers.index('Site Name')+1).value=file.strip('.log')
        
        with open(dir_in+'\\'+file,'r') as f:
            lines=f.readlines()
                    
            for line in lines:
                if "nodeIpAddress" in line:
                    node_ip_addr=line.split(" ")[-1].strip("\n")
                    ws.cell(row=site_count+1,column=headers.index('nodeIpAddress')+1).value=node_ip_addr
                                    
                if "ipAddress" in line:
                    ip_addr=line.split(" ")[-1].strip("\n")
                    ws.cell(row=site_count+1,column=headers.index('ipAddress')+1).value=ip_addr
                                    
                if "Subrack" in line and "vid" in line:
                    ip_if=line.split("IpInterface=")[1][0]
                    vid=line.split(" ")[-1].strip("\n")
                    if ip_if=="1":
                        ws.cell(row=site_count+1,column=headers.index('IpInterface-1')+1).value=vid
                    if ip_if=="2":
                        ws.cell(row=site_count+1,column=headers.index('IpInterface-2')+1).value=vid


    for cell in ws[1]:
        double = Side(border_style="double")
        cell.border=Border(top=double, left=double, right=double, bottom=double)
        if cell.value == "Site Name":
            cell.font=Font(color="00FFFFFF",bold=True)
            cell.fill=PatternFill(start_color="00000000", end_color="00000000",fill_type = "solid")
        if cell.value == "nodeIpAddress":
            cell.font=Font(color="00FFFFFF",bold=True)
            cell.fill=PatternFill(start_color="00660066", end_color="00660066",fill_type = "solid")
        if cell.value == 'ipAddress':
            cell.font=Font(color="00FFFFFF",bold=True)
            cell.fill=PatternFill(start_color="00000080", end_color="00000080",fill_type = "solid")
        if cell.value == 'IpInterface-1':
            cell.font=Font(color="00FFFFFF",bold=True)
            cell.fill=PatternFill(start_color="00008000", end_color="00008000",fill_type = "solid")
        if cell.value == 'IpInterface-2':
            cell.font=Font(color="00FFFFFF",bold=True)
            cell.fill=PatternFill(start_color="000000FF", end_color="000000FF",fill_type = "solid")
    
    # if not os.path.exists(os.path.join(os.getcwd(),'output')):
    #     os.mkdir(os.path.join(os.getcwd(),'output'))
    # wb.save(os.path.join(os.getcwd(),'Output','3G_parsed.xlsx'))
    # #wb.save(r"C:\Users\USER\Desktop\tanvir vai script\3G_(parsed).xlsx")
    # wb.close()

    return wb,site_count


def parser_2G_4G(files):
    headers=['Site Name','RATs',
            'OAM IP Address','Abis IP Address','Iub IP Address','S1_X2 IP Address',
            'OAM Gateway IP','Abis Gateway IP','Iub Gateway IP','S1_X2 Gateway IP',
            'OAM VLAN ID','Abis VLAN ID','Iub VLAN ID','S1_X2 VLAN ID']

    wb=Workbook()
    ws=wb.active

    for i,j in zip(range(1,len(headers)+1),headers):
        ws.cell(row=1,column=i).value=j

    site_count=0
    ip_last_col_next=7
    gw_last_col_next=11
    vlan_last_col_next=15
    
    
    for file in files:
        site_count=site_count+1
        ws.cell(row=site_count+1,column=headers.index('Site Name')+1).value=file.strip('.log')
        search_flag=False

        with open(dir_in+'\\'+file,'r') as f:
            lines=f.readlines()

                    
            for line in lines:

                headers=[cell.value for cell in ws[1]]

                if line.startswith('$rats = '):
                    rats=line.split('$rats = ')[1]
                    ws.cell(row=site_count+1,column=headers.index('RATs')+1).value=rats

                if 'lhget router address' in line:
                    search_flag=True

                if search_flag:

                    if "TwampInitiator" in line and "TwampTestSession" in line:
                        continue
                    
                    if line.startswith("Router=vr_") and 'InterfaceIPv4' in line and 'AddressIPv4' in line:
                        rat_ip_col=line.split("Router=vr_")[1].split(",")[0] + ' IP Address'
                        if rat_ip_col not in headers:
                            ws.insert_cols(ip_last_col_next)
                            ws.cell(row=1,column=ip_last_col_next).value=rat_ip_col
                            ip_last_col_next+=1
                        headers=[cell.value for cell in ws[1]]
                        ws.cell(row=site_count+1,column=headers.index(rat_ip_col)+1).value=line.strip("\n").strip().split(" ")[-1]                    
                            
                    if line.startswith("Router=vr_") and 'RouteTableIPv4Static' in line and 'NextHop' in line:
                        gw_ip_col=line.split("Router=vr_")[1].split(",")[0] + ' Gateway IP'
                        if gw_ip_col not in headers:
                            ws.insert_cols(gw_last_col_next)
                            ws.cell(row=1,column=gw_last_col_next).value=gw_ip_col
                            gw_last_col_next+=1
                        headers=[cell.value for cell in ws[1]]
                        ws.cell(row=site_count+1,column=headers.index(gw_ip_col)+1).value=line.strip("\n").strip().split(" ")[-1]
                            
                            

                    if line.startswith("VlanPort=") and 'vlanId' in line:
                        vlan_col=line.split("VlanPort=")[1].split(" ")[0] + ' VLAN ID'
                        if vlan_col not in headers:
                            ws.insert_cols(vlan_last_col_next)
                            ws.cell(row=1,column=vlan_last_col_next).value=vlan_col
                            vlan_last_col_next+=1
                        headers=[cell.value for cell in ws[1]]
                        ws.cell(row=site_count+1,column=headers.index(vlan_col)+1).value=line.strip("\n").strip().split(" ")[-1]



    for cell in ws[1]:
        double = Side(border_style="double")
        cell.border=Border(top=double, left=double, right=double, bottom=double)
        if cell.value == "Site Name":
            cell.font=Font(color="00FFFFFF",bold=True)
            cell.fill=PatternFill(start_color="00000000", end_color="00000000",fill_type = "solid")
        if cell.value == "RATs":
            cell.font=Font(color="00FFFFFF",bold=True)
            cell.fill=PatternFill(start_color="00660066", end_color="00660066",fill_type = "solid")
        if cell.value.endswith('IP Address'):
            cell.font=Font(color="00FFFFFF",bold=True)
            cell.fill=PatternFill(start_color="00000080", end_color="00000080",fill_type = "solid")
        if cell.value.endswith('Gateway IP'):
            cell.font=Font(color="00FFFFFF",bold=True)
            cell.fill=PatternFill(start_color="00008000", end_color="00008000",fill_type = "solid")
        if cell.value.endswith('VLAN ID'):
            cell.font=Font(color="00FFFFFF",bold=True)
            cell.fill=PatternFill(start_color="000000FF", end_color="000000FF",fill_type = "solid")
        
        
            

    # if not os.path.exists(os.path.join(os.getcwd(),'output')):
    #     os.mkdir(os.path.join(os.getcwd(),'output'))
    # wb.save(os.path.join(os.getcwd(),'Output','2G_4G_parsed.xlsx'))
    # wb.close()

    return wb,site_count

def main():
    init(autoreset=True)
    print(Fore.CYAN+"Welcome to Ericcson Dump Parser script -")
    while True:
        print(Fore.YELLOW+"\n\tOptions available -")
        print("\t\t1. 3G dump parser")
        print("\t\t2. 2G + 4G (and/or 3G) dump parser")
        print("\t\t3. Exit")

        global dir_in

        rat_in=int(input("\n\tPlease select RAT to parse - "))
        while rat_in not in [1,2,3]:
            print("\tWrong input. Please try again.")
            rat_in=input("\n\tPlease select RAT to parse - ")
        
        

        if rat_in==3:
            print("\n\nThank you for using this script. Goodbye.")
            sleep(2)
            sys.exit()
        else:
            dir_in=input("\n\tPlease input the directory : ")
            files=os.listdir(dir_in)
            if rat_in==1:
                print("You selected -  1. 3G dump parser")
                wb,site_count=parser_3G(files)
        
            if rat_in==2:
                print("You selected -  2. 2G + 4G (and/or 3G) dump parser")
                wb,site_count=parser_2G_4G(files)
    
        print(Fore.GREEN+"\n\nTotal {} sites dump parsed.".format(site_count))
        if not os.path.exists(os.path.join(os.getcwd(),'output')):
            os.mkdir(os.path.join(os.getcwd(),'output'))

        filename=dir_in.split("\\")[-1]
        time_now=time_now=datetime.now().strftime(datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))
        file_location=os.path.join(os.getcwd(),'Output',filename+"_"+time_now+"_""_parsed.xlsx")
        wb.save(file_location)
        wb.close()
        print("Output parsed file save in -")
        print(Fore.GREEN+Style.BRIGHT+file_location)
        print("\n\nNew operation -")




try:
    main()
except Exception as e:
    print("Exception occured")
    print(e)

r=input("Please press X to close.")
        
